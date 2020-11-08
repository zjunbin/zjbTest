#!-*-coding: utf-8 -*-
#!Time  : 2020/11/6 15:43
#!@Author : 张俊彬
from openpyxl import load_workbook
from common.readConfig import readConfig
from common.myLog import MyLog


class readExcel:
    def __init__(self, filename, sheetname):
        self.mylog = MyLog('Excel读写')
        self.filename = filename
        self.sheetname = sheetname
        self.read_conf = readConfig().getother('Testconf', 'module')

    def do_excel(self):
        try:
            wb = load_workbook(filename=self.filename)
            self.mylog.debug('====打开Excel====')
        except FileNotFoundError as e:
            self.mylog.error('打开Excel错误{}'.format(e))
            raise e
        try:
            sheet = wb[self.sheetname]
            self.mylog.debug('====打开Sheet====')
        except  KeyError as e:
            self.mylog.error('打开Sheet错误{}'.format(e))
            raise e

        if self.sheetname in self.read_conf:
            list = []
            for i in range(2, sheet.max_row + 1):
                dict = {}
                for j in range(1, sheet.max_column + 1):
                    dict[sheet.cell(1, j).value] = sheet.cell(i, j).value
                list.append(dict)
            return list

    def write_back(self, row, column, value):
        wb = load_workbook(filename=self.filename)
        sheet = wb[self.sheetname]
        sheet.cell(row, column).value = value
        wb.save(self.filename)
        wb.close()
if __name__ == '__main__':
    from common import constans
    file = constans.data_case
    r = readExcel(filename=file,sheetname='register').do_excel()
    print(r)