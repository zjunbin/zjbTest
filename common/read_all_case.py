#!-*-coding: utf-8 -*-
#!Time  : 2020/11/7 19:18
#!@Author : 张俊彬
from openpyxl import load_workbook
from common import constans


class readExcelAll:
    def __init__(self, sheetname):
        self.sheetname = sheetname
        self.filename = constans.data_case
        self.wb = load_workbook(filename=self.filename)


    def DoExcel(self):

        # 读取全部测试用例
        list = {}
        for item in self.wb.sheetnames:
            sheet = self.wb[item]
            dict = []
            for i in range(2, sheet.max_row + 1):
                dict2 = {}
                for j in range(1, sheet.max_column + 1):
                    dict2[sheet.cell(1, j).value] = sheet.cell(i, j).value
                dict.append(dict2)
            list[item] = dict
        # 返回需要使用测试数据
        if self.sheetname in  list.keys():
            data = list[self.sheetname]
            return data

    def write_back(self, row, column, value):
        sheet = self.wb[self.sheetname]
        sheet.cell(row, column).value = value
        self.wb.save(self.filename)
        self.wb.close()
if __name__ == '__main__':
    # data_case = readExcelAll('login').DoExcel()
    # for item in data_case:
    #     print(item)
    pass